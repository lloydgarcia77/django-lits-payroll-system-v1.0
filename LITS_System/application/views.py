from django.shortcuts import render, get_object_or_404, redirect
from django.template.loader import render_to_string
from django.http import JsonResponse, Http404, HttpResponseRedirect, HttpResponse
from django.db.models import Q, Avg, Sum, F
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.views.generic import TemplateView, DetailView
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.urls import reverse, reverse_lazy
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.models import User
from django.db import IntegrityError
from django.core import serializers
from django.utils import timezone
from django.contrib.auth.models import User
# icontains query from a list
# https://thepythonguru.com/python-builtin-functions/reduce/
# https://stackoverflow.com/questions/4824759/django-query-using-contains-each-value-in-a-list
from functools import reduce
# https://docs.python.org/3/library/operator.html
# https://www.geeksforgeeks.org/reduce-in-python/
import operator
from itertools import chain
# for filtering exact salary
from decimal import Decimal

# datetime convertion
import pytz
import datetime
import time
from django.conf import settings
from django.db.models import Value
from django.db.models.functions import Concat
import json
import os

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color, Protection, colors, GradientFill
from openpyxl.utils import get_column_letter
import csv
import xlwt 
from cryptography.fernet import Fernet
import base64
import logging
import traceback
import xlrd 
from django.forms import modelformset_factory, inlineformset_factory
from application.forms import UserAccountCredentialsForm, PersonalForm, MobileNumberForm, SkillsForm, CompanyForm, TelephoneNumberForm, CutOffPeriodForm,AttendanceForm,AttendanceFormManual,EmployeeSalaryForm,EmployeePayrollForm,EmployeeLeavesForm, EmployeeIteneraryForm, EmployeeIteneraryDetailsForm, ConcernsEmployeeForm, ConcernsReplyEmployeeForm
from application.models import PersonalInfo, MobileNumberInfo, SkillsInfo, TelephoneNumberInfo, CompanyInfo, AttendanceInfo, CutOffPeriodInfo, EmployeeSalary, EmployeePayroll,EmployeeLeaves, EmployeeItenerary, EmployeeIteneraryDetails, Concerns, Notifications
from django.forms import modelformset_factory, inlineformset_factory
from datetime import date
# https://www.pythoncircle.com/post/641/encryption-decryption-in-python-django/
# https://www.quora.com/How-do-I-block-inspect-element-on-my-website

category_list = [
    'Uploading File',
    'Delete Uploaded File',
    'Creating Payroll',
    'Updating Payroll Settings',
    'Updating Profile',
    'Updating Password',
    'New User Registration',
    'Deleting Employee',
    'Reply Concerns',
    'Creating Employee Transaction', 
    'Updating Employee Transaction', 
    'Deleting Employee Transaction',
    'Add',
    'Delete',
    ]

level_list =[
        'success',
        'info',
        'warning',
        'error'
        ]

def read_notifications(request,id):
    data = dict()
    current_user = get_object_or_404(User, username=request.user.username)
    notification = get_object_or_404(Notifications, id=id)
    if request.is_ajax():
        if request.method == 'POST':  
            if notification.url:
                data['has_url'] = True
            else: 
                data['has_url'] = False 

            notification.delete()
            
            data['form_is_valid'] = True
            # id = json.loads(request.body)
            # print in readable format
            # print(json.dumps(id, indent=4, sort_keys=True))
            return JsonResponse(data)

    else:
        raise Http404()

def read_all_notifications(request):
    data = dict()
    current_user = get_object_or_404(User, username=request.user.username)
    notifications = Notifications.objects.all().filter(Q(recipient=current_user))
    if request.is_ajax():
        if request.method == 'POST': 
            for notification in notifications:
                notification.delete()
            data['form_is_valid'] = True
            # id = json.loads(request.body)
            # print in readable format
            # print(json.dumps(id, indent=4, sort_keys=True))
            return JsonResponse(data)

    else:
        raise Http404()


def read_attendance_file(attendance_file, request, form, data):

    success =  True # flag for duplicate names on database 

    workbook = xlrd.open_workbook(filename=attendance_file.name, file_contents=attendance_file.read())

    schedule_information = workbook.sheet_by_index(0)
    attendance_stat = workbook.sheet_by_index(1)
    #sheetnames
    sheet_names = workbook.sheet_names()
    print(sheet_names) 
    att_log_report = workbook.sheet_by_name('Att.log report')
    # att_log_report = workbook.sheet_by_index(2)
    # For row 0 and column 0 
    cut_off_period = schedule_information.cell_value(1,1)  

    matching = CutOffPeriodInfo.objects.filter(cut_off_period=cut_off_period).count()
    if not matching > 0:
        #hold the form instance before saving for error checking
        instance = form.save(commit=False)
        instance.attendance_file = attendance_file
        instance.cut_off_period = cut_off_period
        instance.save() #save the form if successful
 

        # dates
        for index in range(3, schedule_information.ncols):
            # Days of Week
            days_of_week = schedule_information.cell_value(2,index)
            days_of_month = schedule_information.cell_value(3,index)

            # print(str(int(days_of_week)),'-', days_of_month, end=" | ")
         
            #checking if db has multiple names might duplicate the attendance record for other matching names
        starting_row = 5
        for row in range(3, att_log_report.nrows):
                starting_row += 2    
                starting_col = 3
                for col in range(att_log_report.ncols):
                    if not starting_row > att_log_report.nrows: 
                        starting_col += 1
                        date = att_log_report.cell_value(3,col)
                        hours = att_log_report.cell_value(starting_row,col)
                        id = att_log_report.cell_value(starting_row-1,2)
                        name = att_log_report.cell_value(starting_row-1,10)   
                        timeIn = hours[:5]
                        timeOut = hours[-5:]    
                        t_diff = None
                        under_time = None
                        if timeIn and timeOut:
                            d_time_in = datetime.datetime.strptime(timeIn, '%H:%M')
                            d_time_out = datetime.datetime.strptime(timeOut, '%H:%M')
                            grace_period =  datetime.datetime.strptime('8:45', '%H:%M')
                            out_time = datetime.datetime.strptime('18:00', '%H:%M')
                            if d_time_in > grace_period:
                                # convert time diff to minutes 1 min = 60 secs
                                t_diff =  int((d_time_in - grace_period).total_seconds() / 60.0)
                            if d_time_out < out_time:
                                under_time = int((out_time - d_time_out).total_seconds() / 60.0)
                        if starting_col < schedule_information.ncols:     
                            days_of_month = schedule_information.cell_value(3, starting_col)    
                            if date != "": 
                                employees = PersonalInfo.objects.all()
                                for employee in employees:
                                    employee_name = '{fname} {mname} {lname}'.format(fname=employee.first_name, mname=employee.middle_name, lname=employee.last_name) 
                                    #print(name,'-',employee_name)
                                    if name in employee_name:
                                        #important notes get will not return when multiple names has been detected on the database or xlsx
                                        #employees = PersonalInfo.objects.annotate(name=Concat('first_name', Value(' '),'middle_name', Value(' '),'last_name'),).filter(Q(name__icontains=name.casefold()))
                                        try:
                                            employees = PersonalInfo.objects.annotate(name=Concat('first_name', Value(' '),'middle_name', Value(' '),'last_name'),).get(Q(name__icontains=name.casefold()))
                                            attendance = AttendanceInfo(employee_profile=employees, cut_off_period=instance, days_of_week=days_of_month, date=str(int(date)), time_in=timeIn, time_out=timeOut, late=t_diff, undertime=under_time)
                                            attendance.save()
                                    
                                        except PersonalInfo.MultipleObjectsReturned:
                                            print('Multiple Names Detected! Cannot be processed!')
                                            
                                            success = False
                                    else:
                                        print('Not Found')
                                # 
                                # attendance = AttendanceInfo(employee_profile=employees, cut_off_period=instance, days_of_week=days_of_month, date=str(int(date)), time_in=timeIn, time_out=timeOut, late=t_diff, undertime=under_time)
                                # attendance.save()

                                # print(row,hours,end=" ")
                                # output = '{row} => ID: [{id}] Day:[{days}] Name: [{name}]- Date: ({date}) - Hours: ({stime}-{etime}) Late: ({dtime}) Under Time: ({utime})'.format(row=row, id=id, days=days_of_month,name=name,date=str(int(date)),stime=timeIn,etime=timeOut, dtime=t_diff, utime=under_time)
                                # print(output, end=" ")
                                # print(row,'=>','[]',str(int(days)), '-(', hours[:5], '-', hours[-5:],')', end=" ")  
                        
                print('')             
        
        

        if success:
            data['success'] = True
        else:
            data['success'] = False
            messages.warning(request, 'Multiple Names Detected! Some employee with duplicate names will not be able to process their attendance!')
        '''
        String -> json.loads() -> json Object
        json Object -> json.dumps() -> String
        '''
        url = reverse_lazy('application:delete_attendance', kwargs={'id':instance.id})
        date_time_str = str(instance.date_created)
        date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d')
        converted_date = date_time_obj.strftime('%b %d, %Y')
        new_record = {
            "id": instance.id,
            "cut_off": cut_off_period,
            "date_created": str(converted_date),
            "attendance_file_path": instance.attendance_file.url,
            "url": str(url),
        }
        #'{"id": "Bob", "languages": ["English", "Fench"]}'
        dict_to_string = json.dumps(new_record)
        data['list'] = json.loads(dict_to_string)
        data['form_is_valid'] = True
        # Applying notifications
        current_user = get_object_or_404(User, username=request.user.username)
        profiles = PersonalInfo.objects.all().distinct().order_by('-id')
        for p in profiles:
            Notifications.objects.create(sender=current_user,recipient=p.fk_user,message="Cut-off attendance was uploaded!",category=category_list[0],level=level_list[1])
             
    else:
        data['success'] = True
        data['form_is_valid'] = False
        messages.error(request, 'Record Already Exists')
   

def encrypt_key(txt):
    try:
        # convert integer etc to string first
        txt = str(txt)
        # get key from settings
        cipher_suite = Fernet(settings.ENCRYPT_KEY)  # key should be byte
        # #input should be byte, so convert the text to byte
        encrypted_text = cipher_suite.encrypt(txt.encode('ascii'))
        # encode to urlsafe base64 format
        encrypted_text = base64.urlsafe_b64encode(
            encrypted_text).decode("ascii")

        return encrypted_text

    except Exception as e:
        # log if an error
        logging.getLogger("error_logger").error(traceback.format_exc())
        return None


def decrypt_key(txt):
    try:
        # base64 decode
        txt = base64.urlsafe_b64decode(txt)
        cipher_suite = Fernet(settings.ENCRYPT_KEY)
        decoded_text = cipher_suite.decrypt(txt).decode("ascii")
        return decoded_text
    except Exception as e:
        # log the error
        logging.getLogger("error_logger").error(traceback.format_exc())
        return None

# Create your views here.


def error_404(request, exception):
    template_name = "error/error_404.html"
    context = {

    }
    return render(request, template_name, context)


def error_500(request):
    template_name = "error/error_500.html"
    context = {

    }
    return render(request, template_name, context)


class PayrollAttendanceIndexPage(LoginRequiredMixin, TemplateView):
    template_name = "index.html"

    def get_context_data(self, *args, **kwargs): 
        context = super(PayrollAttendanceIndexPage, self).get_context_data(*args, **kwargs)
        user = get_object_or_404(User, username=self.request.user.username) 
        if user.is_active and user.is_staff and user.is_superuser: 
            notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
            notifications_count = notifications.count()
            total_emp = PersonalInfo.objects.all().distinct().count()
            #unapproved user registration
            total_unpermitted_emp = User.objects.all().filter(Q(is_active=False) | Q(is_staff=False)).distinct().count()            
            total_active_user_emp = User.objects.all().filter(Q(is_active=True) & Q(is_staff=True)).distinct().count()
            total_attendance = AttendanceInfo.objects.all().distinct().count()  
            total_payroll = EmployeePayroll.objects.all().distinct().count()  
            total_leaves = EmployeeLeaves.objects.all().distinct().count()  
            total_unapproved_leaves = EmployeeLeaves.objects.all().filter(Q(status='Disapproved')).distinct().count()  
            total_approved_leaves = EmployeeLeaves.objects.all().filter(Q(status='Approved')).distinct().count()  
            total_iteneraries = EmployeeItenerary.objects.all().distinct().count()  
            total_unapproved_iteneraries = EmployeeItenerary.objects.all().filter(~Q(checked_by=None) & ~Q(approved_by=None)).distinct().count()  
            total_approved_iteneraries = EmployeeItenerary.objects.all().filter(Q(checked_by=None) & Q(approved_by=None)).distinct().count() 
            total_concern =  Concerns.objects.all().distinct().count() 
            total_unreplied_concern = Concerns.objects.all().filter(Q(reply=None)).distinct().count()   
            emp_payrolls = EmployeePayroll.objects.all()
            year_list = [] 

            yearly_employee_wages_list = []
            yearly_emplpyee_deduction_list = []

            for emp in emp_payrolls:
                year = str(emp.payroll_cutoff_period)[:4]
                if not year in year_list:
                    year_list.append(year)  
            
            #print(year_list)

            for year in year_list:
                total_net_pay = 0
                total_deductions = 0
                cutoffperiod = CutOffPeriodInfo.objects.all().filter(Q(cut_off_period__icontains=year)) 
                for cutoff in cutoffperiod:
                    emp_payroll = EmployeePayroll.objects.all().filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('net_pay')).get('net_pay__sum') 
                    emp_deductions  = EmployeePayroll.objects.all().filter(Q(payroll_cutoff_period=cutoff)).aggregate(Sum('total_deduction')).get('total_deduction__sum') 
                    if emp_payroll:
                        total_net_pay = total_net_pay + emp_payroll 
                        total_deductions = total_deductions + emp_deductions
                    #print(year,'=',cutoff,':',emp_payroll)
                #print('----')
                #print(year,'=',total_net_pay)
                
                yearly_employee_wages_list.append({'year': year, 'value': int(round(total_net_pay,2))})
                yearly_emplpyee_deduction_list.append({'year': year, 'deduction': int(round(total_deductions,2)), 'wages': int(round(total_net_pay,2))})

            dumps_yearly_employee_wages_list = json.dumps(yearly_employee_wages_list)
            loads_yearly_employee_wages_list =json.loads(dumps_yearly_employee_wages_list)
            # print(dumps_yearly_employee_wages_list)
            
            dumps_yearly_employee_deduction_list = json.dumps(yearly_emplpyee_deduction_list)
            loads_yearly_employee_deduction_list =json.loads(dumps_yearly_employee_deduction_list)
                 
            # https://www.educative.io/edpresso/what-is-the-difference-between-jsonloads-and-jsondumps
 
            context.update({
                'user': user, 
                'total_emp': total_emp,
                'total_unpermitted_emp': total_unpermitted_emp,
                'total_active_user_emp': total_active_user_emp,
                'total_attendance': total_attendance,
                'total_payroll': total_payroll,
                'total_leaves': total_leaves,
                'total_unapproved_leaves': total_unapproved_leaves,
                'total_approved_leaves': total_approved_leaves,
                'total_iteneraries': total_iteneraries,
                'total_unapproved_iteneraries': total_unapproved_iteneraries,
                'total_approved_iteneraries': total_approved_iteneraries,
                'total_concern': total_concern,
                'total_unreplied_concern': total_unreplied_concern, 
                'loads_yearly_employee_wages_list': loads_yearly_employee_wages_list,
                'loads_yearly_employee_deduction_list': loads_yearly_employee_deduction_list,
                'notifications': notifications,
                'notifications_count': notifications_count,
            })
            return context
        else:
            raise Http404()


def login_page(request):
    template_name = "registration/login.html"

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')

        user = authenticate(username=username, password=password)

        if user:
            if user.is_active and user.is_staff and user.is_superuser:
                request.session.set_expiry(300)
                login(request, user)
                logged_in_time = get_object_or_404(User, username=user)
                cd = datetime.datetime.now()
                print('==========================', cd)
                cd = timezone.now()
                print('--------------------------', cd)
                logged_in_time.last_login = cd
                logged_in_time.save()
                return HttpResponseRedirect(reverse("index"))
            elif user.is_active and user.is_staff and not user.is_superuser: 
                request.session.set_expiry(300)
                login(request, user)
                logged_in_time = get_object_or_404(User, username=user)
                cd = datetime.datetime.now()
                print('==========================', cd)
                cd = timezone.now()
                print('--------------------------', cd)
                logged_in_time.last_login = cd
                logged_in_time.save()
                return HttpResponseRedirect(reverse("employee_index"))

            else:
                messages.warning(
                    request, "Your account is not active! please contact your administrator.")
        else:
            messages.error(
                request, "Your account is INVALID! please try again.")
    elif request.method == 'GET':
        storage = messages.get_messages(request)
        storage.used = True

    return render(request, template_name)


def register_page(request):
    template_name = "registration/register.html"
    if request.method == 'POST':
        form = UserAccountCredentialsForm(request.POST or None)
        if form.is_valid():
            userForm = form.save(commit=False)
            userForm.is_active = False
            userForm.is_staff = False
            userForm.is_superuser = False
            userForm.save()
            messages.success(
                request, "You have successfully registered, please try to login.")
            return HttpResponseRedirect(reverse("login"))
        else:
            messages.error(request, "Error")
    elif request.method == 'GET':
        form = UserAccountCredentialsForm(request.GET or None)
    context = {
        'form': form,
    }
    return render(request, template_name, context)


@login_required
def user_profile(request):
    # issue
    template_name = "profiles/user_profile.html"
    user = get_object_or_404(User, username=request.user.username)
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    #notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()
    print('-------------->',user)
    # When using in template it was automatically formatted
    # https://ourcodeworld.com/articles/read/555/how-to-format-datetime-objects-in-the-view-and-template-in-django
    last_time_logged = user.last_login
    last_time_joined = user.date_joined
    # last_time_logged = str(user.last_login)[:-6]
    # last_time_joined = str(user.date_joined)[:-6]
    # converted_last_time_logged = datetime.strptime(last_time_logged, '%Y-%m-%d %H:%M:%S.%f')
    # formatted_last_time_logged = datetime.strftime(converted_last_time_logged, '%m/%d/%Y %I:%M:%S %p')

    # converted_last_time_joined = datetime.strptime(last_time_joined, '%Y-%m-%d %H:%M:%S.%f')
    # formatted_last_time_joined = datetime.strftime(converted_last_time_joined, '%m/%d/%Y %I:%M:%S %p')
    # datetime_object = datetime.strptime(str(user.last_login), "%Y-%m-%dT%H:%M:%S.%f%Z")
    # print('-------------------------',datetime_object)
    if user.is_active and user.is_staff and user.is_superuser:
        personalInfoFormset = inlineformset_factory(
            User, PersonalInfo, form=PersonalForm, extra=0, can_delete=False)
        if request.method == "GET":
            piformset = personalInfoFormset(request.GET or None, instance=user)
            record = PersonalInfo.objects.filter(fk_user=user)
            company = CompanyInfo.objects.filter(fk_company_user=user)
            # profile = get_object_or_404(PersonalInfo, fk_user=user)
            # company_details = get_object_or_404(CompanyInfo, fk_company_user=user)
            profiles =  PersonalInfo.objects.all().filter(fk_user=user).distinct()
            company_details = CompanyInfo.objects.all().filter(fk_company_user=user).distinct()
            print(profiles)
        elif request.method == "POST":
            pass

        context = {
            'user': user,
            'last_time_logged': last_time_logged,
            'last_time_joined': last_time_joined,
            'piformset': piformset,
            'record': record,
            'profiles': profiles,
            'company':company,
            'company_details':company_details,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()


@login_required
def new_profile(request):
    data = dict()
    template_name = "profiles/new_profile_personal.html"
    current_user = get_object_or_404(User, username=request.user.username)
    if request.is_ajax():
        if request.method == 'GET':
            form = PersonalForm(request.GET or None)
        elif request.method == 'POST':
            form = PersonalForm(request.POST or None)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.fk_user = current_user
                instance.key_id = encrypt_key(instance.id)
                instance.save() 
                Notifications.objects.create(sender=current_user,recipient=current_user,message="You have been created your new profile!",category=category_list[4],level=level_list[1])
                
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False

        context = {
            'form': form,
        }
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

# On progress
@login_required
def edit_profile(request):
    data = dict()
    template_name = "profiles/edit_profile_personal.html"
    current_user = get_object_or_404(User, username=request.user.username)
    profile = get_object_or_404(PersonalInfo, fk_user=current_user)
    if request.is_ajax():
        if request.method == 'GET':

            form = PersonalForm(request.GET or None, instance=profile)

        elif request.method == 'POST':
            form = PersonalForm(request.POST or None, instance=profile)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.key_id = encrypt_key(profile.id)
                instance.save() 
                Notifications.objects.create(sender=current_user,recipient=current_user,message="You have been updated your profile!",category=category_list[4],level=level_list[1])
                
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False

        context = {
            'form': form,
        }
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()


def save_profile_image(request):
    data = dict()
    current_user = get_object_or_404(User, username=request.user.username)
    if request.is_ajax():
        if request.method == 'POST':
            image = request.FILES['image']
            personal_info, is_created = PersonalInfo.objects.update_or_create(
                fk_user=current_user, defaults={'image': image, })
            if is_created:
                print("Already Exists!")
            else:
                print("New Record!")
            Notifications.objects.create(sender=current_user,recipient=current_user,message="Your profile picture has been updated!",category=category_list[4],level=level_list[1])
            
            data['is_save'] = True
        return JsonResponse(data)
    else:
        raise Http404()


@login_required
def add_mobile_number(request):
    data = dict()
    template_name = "profiles/add_mobile_number.html"
    current_user = get_object_or_404(User, username=request.user.username)
    if request.is_ajax():
        if request.method == 'GET':
            form = MobileNumberForm(request.GET or None)
        elif request.method == 'POST':
            form = MobileNumberForm(request.POST or None)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.fk_mobile_user = current_user
                instance.save()
                Notifications.objects.create(sender=current_user,recipient=current_user,message="New mobile number has been added!",category=category_list[12],level=level_list[1])
            
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False
        context = {
            'form': form,
        }
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()


def delete_mobile_number(request, id):
    data = dict()
    mobile_id = get_object_or_404(MobileNumberInfo, id=id)
    current_user = get_object_or_404(User, username=request.user.username)
    if request.is_ajax():
        if request.method == 'POST':
            print(mobile_id.mobile_number)

            mobile_id.delete()
            Notifications.objects.create(sender=current_user,recipient=current_user,message="Mobile number has been deleted!",category=category_list[13],level=level_list[1])
            
            data['form_is_valid'] = True
            # id = json.loads(request.body)
            # print in readable format
            # print(json.dumps(id, indent=4, sort_keys=True))
            return JsonResponse(data)
    else:
        raise Http404()

def add_skill(request):
    data = dict()
    template_name = "profiles/add_skills.html"
    current_user = get_object_or_404(User, username=request.user.username)
    if request.is_ajax():
        if request.method == 'GET':
            form = SkillsForm(request.GET or None)
        elif request.method == 'POST':
            form = SkillsForm(request.POST or None)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.fk_skills_user = current_user
                instance.save()
                Notifications.objects.create(sender=current_user,recipient=current_user,message="New Skill has been added!",category=category_list[12],level=level_list[1])
            
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False
        context = {
            'form': form,
        }
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

def delete_skill(request, id):
    data = dict()
    skill_id = get_object_or_404(SkillsInfo, id=id)
    current_user = get_object_or_404(User, username=request.user.username)
    if request.is_ajax():
        if request.method == 'POST':
            skill_id.delete()
            data['form_is_valid'] = True
            # id = json.loads(request.body)
            # print in readable format
            # print(json.dumps(id, indent=4, sort_keys=True))
            Notifications.objects.create(sender=current_user,recipient=current_user,message="Skill has been Deleted!",category=category_list[13],level=level_list[1])
            
            return JsonResponse(data)
    else:
        raise Http404()

    return JsonResponse(data)

    # return HttpResponseRedirect(reverse('application:user_profile'))

@login_required
def new_company_details(request):
    data = dict()
    template_name = "profiles/new_company_details.html"
    current_user = get_object_or_404(User, username=request.user.username)

    if request.is_ajax():
        if request.method == 'GET':
            form = CompanyForm(request.GET or None)
        elif request.method == 'POST':
            form = CompanyForm(request.POST or None)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.fk_company_user = current_user 
                instance.save()
                Notifications.objects.create(sender=current_user,recipient=current_user,message="Company profile has been added!",category=category_list[4],level=level_list[1])
            
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False

        context = {
            'form': form,
        }
        data['html_form'] = render_to_string(
            template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def edit_company_details(request):
    data = dict()
    template_name = "profiles/edit_company_details.html"
    current_user = get_object_or_404(User, username=request.user.username)
    company = get_object_or_404(CompanyInfo, fk_company_user=current_user)
    if request.is_ajax():
        if request.method == 'GET':

            form = CompanyForm(request.GET or None, instance=company)

        elif request.method == 'POST':
            form = CompanyForm(request.POST or None, instance=company)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.save()
                Notifications.objects.create(sender=current_user,recipient=current_user,message="Company profile has been updated!",category=category_list[4],level=level_list[1])
            
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False

        context = {
            'form': form,
        }
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def add_telephone_number(request):
    data = dict()
    template_name = "profiles/add_telephone_number.html"
    current_user = get_object_or_404(User, username=request.user.username)
    if request.is_ajax():
        if request.method == 'GET':
            form = TelephoneNumberForm(request.GET or None)
        elif request.method == 'POST':
            form = TelephoneNumberForm(request.POST or None)
            if form.is_valid():
                instance = form.save(commit=False)
                instance.fk_telephone_user = current_user
                instance.save()
                Notifications.objects.create(sender=current_user,recipient=current_user,message="New Telephone number has been added!",category=category_list[12],level=level_list[1])
            
                data['form_is_valid'] = True
            else:
                data['form_is_valid'] = False
        context = {
            'form': form,
        }
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

def delete_telephone_number(request, id):
    data = dict()
    telephone_id = get_object_or_404(TelephoneNumberInfo, id=id)

    if request.is_ajax():
        if request.method == 'POST': 
            telephone_id.delete()
            data['form_is_valid'] = True
            # id = json.loads(request.body)
            # print in readable format
            # print(json.dumps(id, indent=4, sort_keys=True))
            Notifications.objects.create(sender=current_user,recipient=current_user,message="Telephone number has been Deleted!",category=category_list[13],level=level_list[1])
            
            return JsonResponse(data)
    else:
        raise Http404()

@login_required
def cut_off_page(request):
    template_name = "attendance/cut_off_page.html"
    user = get_object_or_404(User, username=request.user.username)

    if user.is_active and user.is_staff and user.is_superuser: 
        cut_off_period_list = CutOffPeriodInfo.objects.all().order_by('-id').distinct()
        notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
        notifications_count = notifications.count()
 
        context = {
            'user':user, 
            'cut_off_period_list': cut_off_period_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

def delete_attendance(request, id):
    data = dict()
    template_name = 'attendance/delete_attendance.html'
    record = get_object_or_404(CutOffPeriodInfo, id=id)
    if request.is_ajax():
        if request.method == 'GET':
            context = {
                'record': record,
            }
            data['html_form'] = render_to_string(template_name, context, request)
        elif request.method == 'POST':
            cut_off_period_list = CutOffPeriodInfo.objects.all().order_by('-id').distinct()
            record.delete()
            path = record.attendance_file.path
            '''
            path. exists(path) - Returns true if the path is a file, directory, or a valid symlink.
            path. isfile(path) - Returns true if the path is a regular file or a symlink to a file.
            path. isdir(path) - Returns true if the path is a directory or a symlink to a directory
            ''' 
            if os.path.isfile(path):
                os.remove(path)
                print('file path is existing')
            else:
                print('file does not exists')

            context = {
                'cut_off_period_list': cut_off_period_list,
            }
            
            data['form_is_valid'] = True
            data['table_records'] = render_to_string('attendance/cut_off_page_table.html', context) 
            # Applying notifications
            current_user = get_object_or_404(User, username=request.user.username)
            profiles = PersonalInfo.objects.all().distinct().order_by('-id')
            for p in profiles:
                Notifications.objects.create(sender=current_user,recipient=p.fk_user,message="Cut-off attendance was deleted!",category=category_list[1],level=level_list[1])
            
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def employee_list_page(request):
    template_name = "employees/employee_list.html"
    user = get_object_or_404(User, username=request.user.username)
    employees_company = User.objects.all().filter(Q(is_active=True) & Q(is_staff=True) ).exclude(username='lloyd.garcia').distinct()
    employees = PersonalInfo.objects.all().filter(fk_user__in=employees_company).exclude(Q(key_id='') | Q(first_name='') | Q(middle_name='') | Q(last_name='last_name'))
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()
    if user.is_active and user.is_staff and user.is_superuser: 
        
 
        context = {
            'user':user, 
            'employees':employees,
            'employees_company': employees_company,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def delete_employee(request, id):
    data = dict()
    template_name = "employees/delete_employee.html"
    personal_info = get_object_or_404(PersonalInfo, id=id)
    current_user = get_object_or_404(User, profile_to_user=personal_info)
    if request.is_ajax():
        if request.method == 'GET':
            
            context = {
            'personal_info': personal_info,
            'user':user,
            }
            data['html_form'] = render_to_string(template_name, context, request)

        elif request.method == 'POST':
            user.delete()
            data['form_is_valid'] = True 
            profiles = PersonalInfo.objects.all().distinct().order_by('-id')
            for p in profiles:
                Notifications.objects.create(sender=current_user,recipient=p.fk_user,message="An employee was deleted!",category=category_list[7],level=level_list[1])
            
  
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def employee_attendance(request, pk):
    template_name = "employees/employee_attendance_list.html"
    user = get_object_or_404(User, username=request.user.username)
    id = decrypt_key(pk)
    employee = get_object_or_404(PersonalInfo, id=id)
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()
    
    if user.is_active and user.is_staff and user.is_superuser: 
        cut_off_period_list = CutOffPeriodInfo.objects.all().order_by('-id').distinct()
        # attendance = AttendanceInfo.objects.all().filter(Q(employee_profile=employee) & Q(cut_off_period__in=cut_off_period_list)).distinct()
        # print(attendance)
        context = {
            'employee':employee,
            'user':user, 
            'cut_off_period_list': cut_off_period_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def employee_attendance_cutoff(request, pk, id):
    template_name = "employees/employee_cutoff_page.html"
    user = get_object_or_404(User, username=request.user.username)
    cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
    pid = decrypt_key(pk)
    employee = get_object_or_404(PersonalInfo, id=pid) 
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()

    attendance = AttendanceInfo.objects.all().filter(Q(employee_profile=employee) & Q(cut_off_period=cutoff)).distinct()
    #print(attendance,cutoff)

    AttendanceInlineFormset = inlineformset_factory(PersonalInfo, AttendanceInfo, form=AttendanceForm, extra=0, can_delete=False)
    
    if user.is_active and user.is_staff and user.is_superuser:  
        is_available = True
        selected_cut_off = AttendanceInfo.objects.filter(Q(cut_off_period=cutoff) & Q(employee_profile=employee)).order_by("-id")
        
        
        if not selected_cut_off:
            is_available = False             
        
        if request.method == 'GET':
            formset = AttendanceInlineFormset(request.GET or None, instance=employee, queryset=AttendanceInfo.objects.filter(cut_off_period=cutoff).order_by("-id"))
            
        elif request.method == 'POST':
            formset = AttendanceInlineFormset(request.POST or None, instance=employee, queryset=AttendanceInfo.objects.filter(cut_off_period=cutoff).order_by("-id"))
            if formset.is_valid(): 
                #https://docs.djangoproject.com/es/2.1/topics/forms/#looping-over-hidden-and-visible-fields
                """
                If you’re manually laying out a form in a template, as opposed to relying on Django’s default form layout, 
                you might want to treat <input type="hidden"> fields differently from non-hidden fields. 
                For example, because hidden fields don’t display anything, putting error messages «next to» the field 
                could cause confusion for your users – so errors for those fields should be handled differently.

                Django provides two methods on a form that allow you to loop over the hidden and visible fields 
                independently: hidden_fields() and visible_fields(). Here’s a modification of an earlier example that uses 
                these two methods:

                {# Include the hidden fields #}
                {% if form in formset.forms %}
                    {% for hidden in form.hidden_fields %}
                        {{ hidden }}    
                    {% endfor %}
                {% endif %}            
                {# Include the visible fields #}
                {% for field in form.visible_fields %}
                    <div class="fieldWrapper">
                        {{ field.errors }}
                        {{ field.label_tag }} {{ field }}
                    </div>
                {% endfor %}

                This example does not handle any errors in the hidden fields. Usually, an error in a hidden field is a 
                sign of form tampering, since normal form interaction won’t alter them. However, you could easily insert 
                some error displays for those form errors, as well.

                """
                # Hidden id causes error id is required
                # for form in formset.forms: 
                #     for hidden in form.hidden_fields():                    
                #         print(hidden)
                
                # form = formset.save(commit=False)
                # form.save()
                myform = formset.save(commit=False)
                for mf in myform:
                    mf.employee_profile = employee
                    mf.save()
                url = HttpResponseRedirect(reverse_lazy('application:employee_side_cutoff_page', kwargs={'id':cutoff.id}))
                Notifications.objects.create(sender=user,recipient=employee.fk_user,url=url.url,message="Your attendance for cut-off {cutoffperiod} was created!".format(cutoffperiod=cutoff.cut_off_period),category=category_list[2],level=level_list[1])
                return HttpResponseRedirect(reverse_lazy('application:employee_attendance_view', kwargs={'pk':employee.key_id}))
            else:
                messages.error(request, "Form Error")
        context = {
            'formset':formset,
            'is_available':is_available,
            'employee':employee,
            'user':user,  
            'cutoff': cutoff,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def employee_attendance_manual_configuration(request, pk, id):
    template_name = "employees/employee_cutoff_manual_page.html"
    user = get_object_or_404(User, username=request.user.username)
    cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
    pid = decrypt_key(pk)
    employee = get_object_or_404(PersonalInfo, id=pid) 
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()

    AttendanceInlineFormset = inlineformset_factory(CutOffPeriodInfo, AttendanceInfo, form=AttendanceFormManual, extra=16, can_delete=False)
    
    if user.is_active and user.is_staff and user.is_superuser:
        if request.method == 'GET':
            formset = AttendanceInlineFormset(request.GET or None, instance=cutoff, queryset=AttendanceInfo.objects.filter(employee_profile=employee).order_by("-id"))
        
        elif request.method == 'POST':
            formset = AttendanceInlineFormset(request.POST or None, instance=cutoff, queryset=AttendanceInfo.objects.filter(employee_profile=employee).order_by("-id"))
            if formset.is_valid():
                myform = formset.save(commit=False)
                for mf in myform:
                    mf.employee_profile = employee
                    mf.cut_off_period = cutoff
                    mf.save()
                url = HttpResponseRedirect(reverse_lazy('application:employee_side_cutoff_page', kwargs={'id':cutoff.id}))
                Notifications.objects.create(sender=user,recipient=employee.fk_user,url=url.url,message="Your attendance for cut-off {cutoffperiod} was manually created!".format(cutoffperiod=cutoff.cut_off_period),category=category_list[2],level=level_list[1])
                return HttpResponseRedirect(reverse_lazy('application:employee_attendance_view', kwargs={'pk':employee.key_id}))
             
            else:
                messages.error(request, "Form Error")
        
        context = {
            'formset':formset,
            'employee':employee,
            'user':user,  
            'cutoff': cutoff,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }
        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def employee_payroll_page(request):
    template_name = "payroll/employee_manage_payroll_page.html"
    user = get_object_or_404(User, username=request.user.username)
    cut_off_period_list = CutOffPeriodInfo.objects.all().order_by('-id').distinct()
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()
    if user.is_active and user.is_staff and user.is_superuser: 
 
        context = {
            'user':user, 
            'cut_off_period_list':cut_off_period_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def employee_payroll_employee_list(request, id):
    template_name = "payroll/employee_manage_payroll_employee_list.html"
    user = get_object_or_404(User, username=request.user.username)
    employees_company = User.objects.all().filter(Q(is_active=True) & Q(is_staff=True) ).exclude(username='lloyd.garcia').distinct()
    employees = PersonalInfo.objects.all().filter(fk_user__in=employees_company).exclude(Q(key_id='') | Q(first_name='') | Q(middle_name='') | Q(last_name='last_name'))
    cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()

     
    if user.is_active and user.is_staff and user.is_superuser: 
 
        context = {
            'user':user, 
            'cutoff': cutoff,
            'employees':employees,
            'employees_company': employees_company,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()
 
def sss_contribution(monthy_salary_credit):
    #https://sssinquiries.com/contributions/new-sss-contributions-table-and-payment-schedule-2019/
    if monthy_salary_credit < 2250:
        return 80
    elif monthy_salary_credit >= 2250 and monthy_salary_credit <= 2749.99:
        return 100
    elif monthy_salary_credit >= 2750 and monthy_salary_credit <= 3249.99:
        return 120
    elif monthy_salary_credit >= 3250 and monthy_salary_credit <= 3749.99:
        return 140
    elif monthy_salary_credit >= 3750 and monthy_salary_credit <= 4249.99:
        return 160
    elif monthy_salary_credit >= 4250 and monthy_salary_credit <= 4749.99:
        return 180
    elif monthy_salary_credit >= 4750 and monthy_salary_credit <= 5249.99:
        return 200
    elif monthy_salary_credit >= 5250 and monthy_salary_credit <= 5749.99:
        return 220
    elif monthy_salary_credit >= 5750 and monthy_salary_credit <= 6249.99:
        return 240
    elif monthy_salary_credit >= 6250 and monthy_salary_credit <= 6749.99:
        return 260
    elif monthy_salary_credit >= 6750 and monthy_salary_credit <= 7249.99:
        return 280
    elif monthy_salary_credit >= 7250 and monthy_salary_credit <= 7749.99:
        return 300
    elif monthy_salary_credit >= 7750 and monthy_salary_credit <= 8249.99:
        return 320
    elif monthy_salary_credit >= 8250 and monthy_salary_credit <= 8749.99:
        return 340
    elif monthy_salary_credit >= 8750 and monthy_salary_credit <= 9249.99:
        return 360
    elif monthy_salary_credit >= 9250 and monthy_salary_credit <= 9749.99:
        return 380
    elif monthy_salary_credit >= 9750 and monthy_salary_credit <= 10249.99:
        return 400
    elif monthy_salary_credit >= 10250 and monthy_salary_credit <= 10749.99:
        return 420
    elif monthy_salary_credit >= 10760 and monthy_salary_credit <= 11249.99:
        return 440
    elif monthy_salary_credit >= 11250 and monthy_salary_credit <= 11749.99:
        return 460
    elif monthy_salary_credit >= 11750 and monthy_salary_credit <= 12249.99:
        return 480
    elif monthy_salary_credit >= 12250 and monthy_salary_credit <= 12749.99:
        return 500
    elif monthy_salary_credit >= 12760 and monthy_salary_credit <= 13249.99:
        return 520
    elif monthy_salary_credit >= 13250 and monthy_salary_credit <= 13749.99:
        return 540
    elif monthy_salary_credit >= 13750 and monthy_salary_credit <= 14249.99:
        return 560
    elif monthy_salary_credit >= 14250 and monthy_salary_credit <= 14749.99:
        return 580
    elif monthy_salary_credit >= 14750 and monthy_salary_credit <= 15249.99:
        return 600
    elif monthy_salary_credit >= 15250 and monthy_salary_credit <= 15749.99:
        return 620
    elif monthy_salary_credit >= 15750 and monthy_salary_credit <= 16249.99:
        return 640
    elif monthy_salary_credit >= 16250 and monthy_salary_credit <= 16749.99:
        return 660
    elif monthy_salary_credit >= 16750 and monthy_salary_credit <= 17249.99:
        return 680
    elif monthy_salary_credit >= 17250 and monthy_salary_credit <= 17749.99:
        return 700
    elif monthy_salary_credit >= 17750 and monthy_salary_credit <= 18249.99:
        return 720
    elif monthy_salary_credit >= 18250 and monthy_salary_credit <= 18749.99:
        return 740
    elif monthy_salary_credit >= 18750 and monthy_salary_credit <= 19249.99:
        return 760
    elif monthy_salary_credit >= 19250 and monthy_salary_credit <= 19749.99:
        return 780
    elif monthy_salary_credit >= 19750:
        return 800 

def day_shift_payroll_computation(data):
    WORKING_DAYS = [
        "MON","TUE","WED","THU","FRI", 
    ]
    REST_DAYS = [
       "SAT", "SUN",
    ]
    LEAVES = [
        'Vacation Leave',
        'Sick Leave',
        'Paternity Leave',
        'Leave without pay',
        'Others',
    ]
    monthly_rate = data["employee_salary"]
    #daily_rate = (monthly_rate * 12) / 313
    daily_rate = (monthly_rate * 12) / 261 #because based on the payroll 5 days a week is counting
    hourly_rate = daily_rate / 8
    min_rate = hourly_rate / 60

    daily_rate = round(daily_rate, 2)
    hourly_rate = round(hourly_rate, 2)
    min_rate = round(min_rate, 2)

    payroll_info = {
        "Basic Pay": 0,
        "Overtime Pay": 0,
        "Legal Holiday": 0,
        "Sunday/Special Holiday": 0,
        "Late/Absences": 0,
        "Gross Pay":0, 
    }

    PCO = {
        "Ordinary Days" : 1.25,
        "Rest Day, Special Day, or Regular Day": 1.30, 
    }

    LEGAL_HOLIDAY = {
        'Regular Holiday': 2.0,
        'Regular Holiday and Rest Day': 2.60,
    }

    SPECIAL_HOLIDAY = {
        'Special Day': 1.30,
        'Special Day and Rest Day': 1.50,
    }

    HOLIDAYS = [
        "Regular Holiday",
        "Special Day",
        "Special Day and Rest Day",
        "Regular Holiday and Rest Day",
    ]
 
    
    if data["Day"] in WORKING_DAYS:
        if data["PCFWD"] in LEAVES:
            if data["PCFWD"] == "Leave without pay":
                payroll_info.update({"Late/Absences": float(daily_rate)})
                return payroll_info 
            else:
                #excempted
                payroll_info.update({"Basic Pay": float(daily_rate)})
                return payroll_info
        
        else: 
            if not data["TimeIn"] or not data["TimeOut"] or data["TimeIn"] == "0:00" or data["TimeOut"] == "0:00" : 
                if data["PCFWD"] in HOLIDAYS:
                    payroll_info.update({"Basic Pay": float(daily_rate)})
                    return payroll_info
                else:
                    
                    payroll_info.update({"Late/Absences": float(daily_rate)})
                    return payroll_info

            else:
                total_late_absences = 0.0
                if data["Late"] or data["Late"] is not None:
                    total_late_absences = total_late_absences + float(data["Late"])
                if data["Undertime"] or data["Undertime"] is not None: 
                    total_late_absences = total_late_absences + float(data["Undertime"]) 
                    total_late_absences = total_late_absences * float(min_rate)

                payroll_info.update({"Late/Absences": total_late_absences})

                if data["Overtime"]: 
                    total_overtime_pay = float(data["Overtime"]) *  PCO[data["PCO"]]
                    payroll_info.update({"Overtime Pay": total_overtime_pay})
                
                if data["PCFWD"] in LEGAL_HOLIDAY: 
                    legal_holiday = float(daily_rate) * float(LEGAL_HOLIDAY[data["PCFWD"]])
                    payroll_info.update({"Legal Holiday": legal_holiday})
                elif data["PCFWD"] in SPECIAL_HOLIDAY: 
                    special_holiday = float(daily_rate) * float(SPECIAL_HOLIDAY[data["PCFWD"]])
                    payroll_info.update({"Sunday/Special Holiday": special_holiday})
                else:
                    payroll_info.update({"Basic Pay": float(daily_rate)}) 
                return payroll_info 

    elif data["Day"] in REST_DAYS:
        if not data["TimeIn"] or not data["TimeOut"] or data["TimeIn"] == "0:00" or data["TimeOut"] == "0:00":
            payroll_info.update({"Basic Pay": 0})
            return payroll_info
        else:
                total_late_absences = 0.0
                if data["Late"] or data["Late"] is not None:
                    total_late_absences = total_late_absences + float(data["Late"])
                if data["Undertime"] or data["Undertime"] is not None: 
                    total_late_absences = total_late_absences + float(data["Undertime"]) 
                    total_late_absences = total_late_absences * float(min_rate)

                payroll_info.update({"Late/Absences": total_late_absences})

                if data["Overtime"]: 
                    total_overtime_pay = float(data["Overtime"]) *  PCO[data["PCO"]]
                    payroll_info.update({"Overtime Pay": total_overtime_pay})
                
                if data["PCFWD"] in LEGAL_HOLIDAY: 
                    legal_holiday = float(daily_rate) * float(LEGAL_HOLIDAY[data["PCFWD"]])
                    payroll_info.update({"Legal Holiday": legal_holiday})
                elif data["PCFWD"] in SPECIAL_HOLIDAY: 
                    special_holiday = float(daily_rate) * float(SPECIAL_HOLIDAY[data["PCFWD"]])
                    payroll_info.update({"Sunday/Special Holiday": special_holiday})
                else:
                    payroll_info.update({"Basic Pay": float(daily_rate)}) 
                return payroll_info 


@login_required
def employee_create_payroll(request, key, id): 
    template_name = "payroll/employee_create_payroll.html"
    user = get_object_or_404(User, username=request.user.username)
    cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
    pid = decrypt_key(key)
    employee = get_object_or_404(PersonalInfo, id=pid) 
    company_details = get_object_or_404(User, username=employee.fk_user.username)
    employee_salary = get_object_or_404(EmployeeSalary, employee_salary_fk=employee)
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()

    #https://www.programiz.com/python-programming/datetime/current-datetime
    today = date.today()
    date_today = today.strftime("%m/%d/%Y") 
    attendance = AttendanceInfo.objects.all().filter(Q(employee_profile=employee) & Q(cut_off_period=cutoff)).order_by("-id").distinct()
    
    #check if payroll exists
    emp_payroll_count = EmployeePayroll.objects.all().filter(Q(employee_fk=employee) & Q(payroll_cutoff_period=cutoff)).distinct().count()
    
    basic_pay = 0
    overtime_pay = 0
    legal_holiday = 0
    sunday_spec_holiday = 0
    late_absences = 0

    for a in attendance:

        attendance_data = {
            "Day": a.days_of_week,
            "TimeIn": a.time_in,
            "TimeOut": a.time_out,
            "PCFWD": a.payment_computation_for_work_done,
            "Late":a.late,
            "Undertime": a.undertime,
            "Overtime": a.overtime,
            "PCO": a.payment_computation_overtime,
            "employee_salary":employee_salary.amount,
        } 

        output = day_shift_payroll_computation(attendance_data)

        basic_pay = basic_pay + output["Basic Pay"]
        overtime_pay = overtime_pay + output["Overtime Pay"]
        legal_holiday = legal_holiday + output["Legal Holiday"]
        sunday_spec_holiday = sunday_spec_holiday + output["Sunday/Special Holiday"]
        late_absences = late_absences + output["Late/Absences"]
        # print(attendance_data)
        # print(output)
        # print("")
    
    # print("Total Basic Pay:",round(basic_pay, 2))
    # print("Total Overtime Pay:",round(overtime_pay, 2))
    # print("Total Legal Holiday:",round(legal_holiday, 2))
    # print("Total Sunday/Special Holiday:",round(sunday_spec_holiday, 2))
    # print("Total Late/Absences:",round(late_absences, 2))
    # print("Total SSS: ", round(sss_contribution(float(employee_salary.amount))))
    total_gross_pay = basic_pay + overtime_pay + legal_holiday + sunday_spec_holiday + float(employee_salary.allowance)
    total_deduction = sss_contribution(float(employee_salary.amount)) + late_absences
    total_net_pay = total_gross_pay - total_deduction
    # print("Total Gross pay: ", total_gross_pay)
    # print("Total Deduction: ", total_deduction)
    # print("Total Net pay: ", total_net_pay) 
        #break
        # if a.time_in and a.time_out:
        #     time_format = '%H:%M'
        #     time_delta = (datetime.datetime.strptime(a.time_out,time_format) - datetime.datetime.strptime(a.time_in,time_format)).total_seconds()
        #     time_diff_min = time_delta / 60
        #     total_min_render = total_min_render + time_diff_min
        #     display = "Day: {day} | Time In:{timein} | Time Out: {timeout} | [Time Diff: {tdiff}]| Late: {late} | undertime: {undertime} | overtime: {overtime}".format(day=a.days_of_week,timein=a.time_in,timeout=a.time_out,late=a.late,undertime=a.undertime,overtime=a.overtime, tdiff=time_diff_min)
 

    if user.is_active and user.is_staff and user.is_superuser: 
        if request.method == 'GET':
             #check if there is an existing record  
            if emp_payroll_count > 0: 
                #raise Http404()
                return HttpResponseRedirect(reverse_lazy('application:employee_edit_payroll', kwargs={'key':employee.key_id,'id': cutoff.id}))
        
        elif request.method == 'POST': 
            _basicPay = request.POST.get('_basicPay',"0") if request.POST.get('_basicPay',"0")  != "" else "0"
            _allowance = request.POST.get('_allowance',"0") if request.POST.get('_allowance',"0")  != "" else "0"
            _overtimePay = request.POST.get('_overtimePay',"0") if request.POST.get('_overtimePay',"0")  != "" else "0"
            _legalHoliday = request.POST.get('_legalHoliday',"0") if request.POST.get('_legalHoliday',"0")  != "" else "0"
            _sundaySpecialHoliday = request.POST.get('_sundaySpecialHoliday',"0") if request.POST.get('_sundaySpecialHoliday',"0")  != "" else "0"
            _lateAbsences = request.POST.get('_lateAbsences',"0") if request.POST.get('_lateAbsences',"0")  != "" else "0"
            _salaryCashAdvance = request.POST.get('_salaryCashAdvance',"0") if request.POST.get('_salaryCashAdvance',"0")  != "" else "0"
            _grossPay = request.POST.get('_grossPay',"0") if request.POST.get('_grossPay',"0")  != "" else "0"
            _netPay = request.POST.get('_netPay',"0") if request.POST.get('_netPay',"0")  != "" else "0" 
            _philhealContribution = request.POST.get('_philhealContribution',"0") if request.POST.get('_philhealContribution',"0")  != "" else "0"
            _pagibigContribution = request.POST.get('_pagibigContribution',"0") if request.POST.get('_pagibigContribution',"0")  != "" else "0"
            _sssPremius = request.POST.get('_sssPremius',"0") if request.POST.get('_sssPremius',"0")  != "" else "0"
            _withholdingTax = request.POST.get('_withholdingTax',"0") if request.POST.get('_withholdingTax',"0")  != "" else "0"
            _pagibigLoan = request.POST.get('_pagibigLoan',"0") if request.POST.get('_pagibigLoan',"0")  != "" else "0"
            _deductionSalaryCashAdvance = request.POST.get('_deductionSalaryCashAdvance',"0") if request.POST.get('_deductionSalaryCashAdvance',"0")  != "" else "0"
            _totalDeduction = request.POST.get('_totalDeduction',"0") if request.POST.get('_totalDeduction',"0")  != "" else "0"
            _cboSSS = request.POST.get('_cboSSS',"0") 

            sss = _sssPremius if _cboSSS else 0
          
            print('SCA:', _salaryCashAdvance)
            print('PH',_philhealContribution, '-',request.POST.get('_philhealContribution',"0"))
            print('PC',_pagibigContribution)
            print('WT',_withholdingTax)
            print('PL',_pagibigLoan)
            print('SCA',_deductionSalaryCashAdvance)
            

            #employee_payroll_manage_employee_list
            new, existing = EmployeePayroll.objects.update_or_create(employee_fk=employee, payroll_cutoff_period=cutoff, 
            defaults= {
                #'payroll_date': date_today, 
                'monthly_rate':float(employee_salary.amount),  
                'monthly_allowance': 0, # for future purposes 
                'basic_pay': round(basic_pay, 2),
                'allowance': float(employee_salary.allowance),
                'overtime_pay': round(overtime_pay, 2),
                'legal_holiday': round(legal_holiday, 2),
                'special_holiday':  round(sunday_spec_holiday, 2),
                'late_or_absences': round(late_absences, 2),
                'salary_or_cash_advance': _salaryCashAdvance,
                'gross_pay': _grossPay,
                'sss_premiums': sss,
                'philhealth_contribution': _philhealContribution,
                'pagibig_contribution': _pagibigContribution,
                'withholding_tax': _withholdingTax,
                'pagibig_loan': _pagibigLoan,
                'deducted_salary_cash_advance': _deductionSalaryCashAdvance,
                'total_deduction': _totalDeduction,
                'net_pay': _netPay, 
                })
            url = HttpResponseRedirect(reverse_lazy('application:employee_side_view_payroll_page', kwargs={'id':cutoff.id}))
            Notifications.objects.create(sender=user,recipient=employee.fk_user,url=url.url,message="Your payroll for cut-off {cutoffperiod} was created!".format(cutoffperiod=cutoff.cut_off_period),category=category_list[2],level=level_list[1])
            return HttpResponseRedirect(reverse_lazy('application:employee_payroll_manage_employee_list', kwargs={'id': cutoff.id}))
        
        context = {
            'user':user, 
            'cutoff':cutoff,
            'employee':employee,
            'date_today':date_today,
            'company_details': company_details,
            'attendance':attendance,
            'employee_salary':employee_salary,

            'basic_pay': round(basic_pay, 2),
            'overtime_pay': round(overtime_pay, 2),
            'legal_holiday': round(legal_holiday, 2),
            'sunday_special_holiday': round(sunday_spec_holiday, 2),

            'late_absences': round(late_absences, 2),
            'sss_contribution': round(sss_contribution(float(employee_salary.amount))),

            'gross_pay': round(total_gross_pay, 2),
            'total_deduction': round(total_deduction, 2),
            'net_pay': round(total_net_pay, 2),

            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def employee_edit_payroll(request, key, id): 
    template_name = "payroll/employee_edit_payroll.html"
    user = get_object_or_404(User, username=request.user.username)
    cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
    pid = decrypt_key(key)
    employee = get_object_or_404(PersonalInfo, id=pid)  
    employee_salary = get_object_or_404(EmployeeSalary, employee_salary_fk=employee)
    company_details = get_object_or_404(User, username=employee.fk_user.username)
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()

    today = date.today()
    date_today = today.strftime("%m/%d/%Y") 

    attendance = AttendanceInfo.objects.all().filter(Q(employee_profile=employee) & Q(cut_off_period=cutoff)).order_by("-id").distinct()
    
    #check if payroll exists
    emp_payroll_count = EmployeePayroll.objects.all().filter(Q(employee_fk=employee) & Q(payroll_cutoff_period=cutoff)).distinct().count()
    

    if user.is_active and user.is_staff and user.is_superuser: 
        if request.method == 'GET':
              
            if emp_payroll_count <= 0: 
                return HttpResponseRedirect(reverse_lazy('application:employee_create_payroll', kwargs={'key':employee.key_id,'id': cutoff.id}))
            else: 
                emp_payroll = get_object_or_404(EmployeePayroll, employee_fk=employee, payroll_cutoff_period=cutoff)
                forms = EmployeePayrollForm(request.GET or None, instance=emp_payroll) 

        elif request.method == 'POST':
            emp_payroll = get_object_or_404(EmployeePayroll, employee_fk=employee, payroll_cutoff_period=cutoff)
            forms = EmployeePayrollForm(request.POST or None, instance=emp_payroll) 
            if forms.is_valid():
                instance = forms.save(commit=False)
                instance.save() 
                url = HttpResponseRedirect(reverse_lazy('application:employee_side_view_payroll_page', kwargs={'id':cutoff.id}))
                Notifications.objects.create(sender=user,recipient=employee.fk_user,url=url.url,message="Your payroll for cut-off {cutoffperiod} was updated!".format(cutoffperiod=cutoff.cut_off_period),category=category_list[2],level=level_list[1])               
                return HttpResponseRedirect(reverse_lazy('application:employee_payroll_manage_employee_list', kwargs={'id': cutoff.id}))
            else:
                print("Form Error!")
        context = {
            'user':user, 
            'cutoff':cutoff,
            'employee':employee,
            'date_today':date_today,
            'employee_salary': employee_salary,
            'attendance': attendance,
            'company_details':company_details,
            'forms': forms, 
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def employee_delete_payroll(request, key, id):
    data = dict()
    template_name = "payroll/employee_delete_payroll.html"
    user = get_object_or_404(User, username=request.user.username)
    cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
    pid = decrypt_key(key)
    employee = get_object_or_404(PersonalInfo, id=pid)  
    emp_payroll = get_object_or_404(EmployeePayroll, employee_fk=employee, payroll_cutoff_period=cutoff)
    link = HttpResponseRedirect(reverse_lazy('application:employee_payroll_manage_employee_list', kwargs={'id': cutoff.id}))
    if request.is_ajax():
        if request.method == 'GET':
            pass
        elif request.method == 'POST':
            emp_payroll.delete()
            data['form_is_valid'] = True
            url = HttpResponseRedirect(reverse_lazy('application:employee_side_view_payroll_page', kwargs={'id':cutoff.id}))
            Notifications.objects.create(sender=user,recipient=employee.fk_user,url=url.url,message="Your payroll for cut-off {cutoffperiod} was deleted!".format(cutoffperiod=cutoff.cut_off_period),category=category_list[1],level=level_list[1])               
         
        context = {
            'user':user, 
            'cutoff':cutoff,
            'employee':employee,
            'emp_payroll': emp_payroll, 
        }
        data['link'] = link.url
        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def employee_manage_leaves(request):
    template_name = "leaves/employee_leaves_page.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()
    leaves = EmployeeLeaves.objects.all().filter().order_by('-id').distinct()

    if user.is_active and user.is_staff and user.is_superuser: 
        if request.method == 'GET':
            pass
        elif request.method == 'POST':
            pass

        context = {
            'user':user,
            'employee': employee,
            'leaves': leaves,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }
        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def employee_view_employee_leaves(request, id):
    template_name="leaves/employee_view_leave_page.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    leave = get_object_or_404(EmployeeLeaves, id=id)
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()

    employee_list = PersonalInfo.objects.all().order_by('-id').distinct()

    status_list = [
        'Disapproved',
        'Approved',
    ]

 
    if user.is_active and user.is_staff and user.is_superuser: 
        if request.method == 'GET': 
            pass
        elif request.method == 'POST':
            _leaveCredits = request.POST.get('_leaveCredits', "0")
            _lessThisApplication = request.POST.get('_lessThisApplication', "0")
            _balanceAsOfThisDate = request.POST.get('_balanceAsOfThisDate', "0")
            _status = request.POST['statusList']
            _notedBy = request.POST['notedByList']
            _checkedBy = request.POST['checkedByList']
            _approvedBy = request.POST['approvedByList']
 

            new, existing = EmployeeLeaves.objects.update_or_create(id=id, defaults={
                'status': _status,
                'leave_credits': _leaveCredits,
                'less_this_application': _lessThisApplication,
                'balance_as_of_this_date': _balanceAsOfThisDate,
                'noted_by': _notedBy,
                'checked_by': _checkedBy,
                'approved_by': _approvedBy,

            })
            url = HttpResponseRedirect(reverse_lazy('application:employee_side_view_leave_form', kwargs={'id':leave.id}))
            Notifications.objects.create(sender=user,recipient=employee.fk_user,url=url.url,message="Your leave form was updated by the administrator!",category=category_list[10],level=level_list[1])               
            return HttpResponseRedirect(reverse_lazy('application:employee_manage_leaves_page'))
        context = {
            'user':user,
            'employee': employee,
            'leave': leave, 
            'employee_list': employee_list,
            'status_list': status_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }
        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def employee_manage_itenerary(request):
    template_name = "iteneraries/employee_manage_iteneraries_page.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)    
    itenerary_list = EmployeeItenerary.objects.all().order_by('-id').distinct()
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()
    if user.is_active and user.is_staff and user.is_superuser: 
        if request.method == 'GET':
            pass
        elif request.method == 'POST':
            pass
        context = {
            'user':user,
            'employee': employee, 
            'itenerary_list': itenerary_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }
        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def employee_view_itenerary(request, id):
    template_name = "iteneraries/employee_view_iteneraries_page.html"
    user = get_object_or_404(User, username=request.user.username)
    itenerary = get_object_or_404(EmployeeItenerary, id=id)
    employee_list = PersonalInfo.objects.all().order_by('-id').distinct()
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()


    if user.is_active and user.is_staff and user.is_superuser: 
        if request.method == 'GET':
            pass
        elif request.method == 'POST':
            _notedBy = request.POST['notedByList']
            _checkedBy = request.POST['checkedByList']
            _approvedBy = request.POST['approvedByList']
 

            new, existing = EmployeeItenerary.objects.update_or_create(id=id, defaults={ 
                'noted_by': _notedBy,
                'checked_by': _checkedBy,
                'approved_by': _approvedBy,

            })
            url = HttpResponseRedirect(reverse_lazy('application:employee_side_view_iteneraries_page', kwargs={'id':itenerary.id}))
            Notifications.objects.create(sender=user,recipient=itenerary.employee_itenerary_fk.fk_user,url=url.url,message="Your itinerary form was updated by the administrator!",category=category_list[10],level=level_list[1])               
            return HttpResponseRedirect(reverse_lazy('application:employee_manage_iteneraries_page'))
        context = {
            'user':user, 
            'itenerary': itenerary,
            'employee_list': employee_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }
        return render(request, template_name, context)
    else:
        raise Http404()


@login_required
def employee_manage_concerns(request):
    template_name = "concerns/employee_manage_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    concern_list = Concerns.objects.all().filter(Q(sender=employee)).order_by('-id').distinct()
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()
    if user.is_active and user.is_staff and user.is_superuser: 
        if request.method == 'GET':
            pass
        elif request.method == 'POST':
            pass
        context = {
            'user':user,
            'employee': employee, 
            'concern_list': concern_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }
        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def employee_create_concern(request):
    data = dict()
    template_name = "concerns/employee_create_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)

    if user.is_active and user.is_staff and  user.is_superuser: 
        if request.method == 'GET':
            form = ConcernsEmployeeForm(request.GET or None)
        elif request.method == 'POST':
            form = ConcernsEmployeeForm(request.POST or None)   
            if form.is_valid():
                instance = form.save(commit=False)         
                instance.sender  = employee
                instance.save()
                #notification
                if instance.receiver.fk_user.is_superuser:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_manage_inbox_concerns_page'))
                else:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_inbox_concerns_page'))
                Notifications.objects.create(sender=user,recipient=instance.receiver.fk_user,url=url.url,message="You had received new concern from {sender}".format(sender=employee),category=category_list[9],level=level_list[1])               
                #end notificaiton
                date_time_str = str(instance.date_filed)
                date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d')
                converted_date = date_time_obj.strftime('%b %d, %Y')
                edit_url = reverse_lazy('application:employee_edit_concerns_page', kwargs={'id':instance.id})
                delete_url = reverse_lazy('application:employee_delete_concerns_page', kwargs={'id':instance.id})
                view_url = reverse_lazy('application:employee_view_concerns_page', kwargs={'id':instance.id})
                concern_date_dict = {
                    "id": instance.id,
                    "subject": instance.subject,
                    "date_filed": str(converted_date),
                    "edit_url": str(edit_url),
                    "view_url": str(view_url),
                    "delete_url": str(delete_url),
                }
                dict_to_string = json.dumps(concern_date_dict)
                data['concern_date_dict'] = json.loads(dict_to_string)
                data['form_is_valid'] = True

        context = {
            'form': form,
            'user': user,  
            'employee': employee,  
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def employee_edit_concern(request, id):
    data = dict()
    template_name = "concerns/employee_edit_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    concern = get_object_or_404(Concerns, id=id)

    if user.is_active and user.is_staff and  user.is_superuser: 
        if request.method == 'GET':
            form = ConcernsEmployeeForm(request.GET or None, instance=concern)
        elif request.method == 'POST':
            form = ConcernsEmployeeForm(request.POST or None, instance=concern)   
            if form.is_valid():
                instance = form.save(commit=False)         
                instance.sender  = employee
                instance.save()

                if instance.receiver.fk_user.is_superuser:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_manage_inbox_concerns_page'))
                else:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_inbox_concerns_page'))
                Notifications.objects.create(sender=user,recipient=instance.receiver.fk_user,url=url.url,message="Concern from {sender} was updated".format(sender=employee),category=category_list[10],level=level_list[1])               

                date_time_str = str(instance.date_filed)
                date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d')
                converted_date = date_time_obj.strftime('%b %d, %Y')
                edit_url = reverse_lazy('application:employee_edit_concerns_page', kwargs={'id':instance.id})
                delete_url = reverse_lazy('application:employee_delete_concerns_page', kwargs={'id':instance.id})
                view_url = reverse_lazy('application:employee_view_concerns_page', kwargs={'id':instance.id})
                concern_date_dict = {
                    "id": instance.id,
                    "subject": instance.subject,
                    "date_filed": str(converted_date),
                    "edit_url": str(edit_url),
                    "view_url": str(view_url),
                    "delete_url": str(delete_url),
                }
                dict_to_string = json.dumps(concern_date_dict)
                data['concern_date_dict'] = json.loads(dict_to_string)
                data['form_is_valid'] = True

        context = {
            'form': form,
            'concern': concern,
            'user': user,  
            'employee': employee,  
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def employee_view_concern(request, id):
    data = dict()
    template_name = "concerns/employee_view_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    concern = get_object_or_404(Concerns, id=id)

    if user.is_active and user.is_staff and user.is_superuser: 
        if request.method == 'GET':
            pass
       
        context = { 
            'user': user,  
            'employee': employee,  
            'concern': concern,
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def employee_delete_concern(request, id):
    data = dict()
    template_name = "concerns/employee_delete_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    concern = get_object_or_404(Concerns, id=id)

    if user.is_active and user.is_staff and user.is_superuser: 
        if request.method == 'GET':
            context = { 
            'user': user,  
            'employee': employee,  
            'concern': concern,
            } 
            data['html_form'] = render_to_string(template_name, context, request)
        elif request.method == 'POST': 
            if concern.receiver.fk_user.is_superuser:
                url = HttpResponseRedirect(reverse_lazy('application:employee_manage_inbox_concerns_page'))
            else:
                url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_inbox_concerns_page')) 
            Notifications.objects.create(sender=user,recipient=concern.receiver.fk_user,url=url.url,message="Your concern from {sender}".format(sender=concern.sender),category=category_list[11],level=level_list[1])               
            concern.delete()
            data['form_is_valid'] = True 
            
        return JsonResponse(data)
    else:
        raise Http404()
 
@login_required
def employee_manage_inbox_concern(request):
    template_name = "concerns/employee_manage_inbox_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()

    concern_list = Concerns.objects.all().filter(Q(receiver=employee)).order_by('-id').distinct()

    if user.is_active and user.is_staff and user.is_superuser:
        if request.method == 'GET':
            pass 
        context = {
            'user': user,  
            'employee': employee, 
            'concern_list': concern_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()


@login_required
def employee_reply_concern(request, id):
    data = dict()
    template_name = "concerns/employee_reply_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    concern = get_object_or_404(Concerns, id=id)

    if user.is_active and user.is_staff and user.is_superuser: 
        if request.method == 'GET':
            form = ConcernsReplyEmployeeForm(request.GET or None, instance=concern)
        elif request.method == 'POST':
            form = ConcernsReplyEmployeeForm(request.POST or None, instance=concern)   
            if form.is_valid():
                instance = form.save(commit=False)   
                instance.save()  
                if instance.sender.fk_user.is_superuser:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_manage_inbox_concerns_page'))
                else:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_inbox_concerns_page'))
                Notifications.objects.create(sender=user,recipient=instance.sender.fk_user,url=url.url,message="Your concern has been repilied by {receiver}".format(receiver=concern.receiver),category=category_list[8],level=level_list[1])               


                data['form_is_valid'] = True

        context = {
            'form': form,
            'user': user,  
            'employee': employee,  
            'concern': concern,
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()


@login_required
def employee_payroll_base_salary_settings(request, key):
    template_name = "employees/employee_salary_settings.html"
    current_user = get_object_or_404(User, username=request.user.username)
    data = dict()
    pid = decrypt_key(key)
    employee = get_object_or_404(PersonalInfo, id=pid) 
    
    count = EmployeeSalary.objects.all().filter(employee_salary_fk=employee).distinct()
 
    if count: 
        EmployeeSalaryInlineFomrset = inlineformset_factory(PersonalInfo, EmployeeSalary, form=EmployeeSalaryForm, extra=0 , can_delete=False)
    else:
        EmployeeSalaryInlineFomrset = inlineformset_factory(PersonalInfo, EmployeeSalary, form=EmployeeSalaryForm, extra=1 , can_delete=False)

    if request.is_ajax():
        if request.method == 'GET': 
            formset = EmployeeSalaryInlineFomrset(request.GET or None, instance=employee)
        elif request.method == 'POST':
            formset = EmployeeSalaryInlineFomrset(request.POST or None, instance=employee)
            if formset.is_valid():
                myform = formset.save(commit=False)
                for mf in myform:
                    mf.employee_salary_fk = employee
                    mf.save()
                data['form_is_valid'] = True
                Notifications.objects.create(sender=current_user,recipient=employee.fk_user,message="Your payroll settings was updated!",category=category_list[3],level=level_list[1])
            else:
                messages.error(request, "Form Error")


        context = {
            'employee': employee,
            'formset': formset,
        }

        data['html_form'] = render_to_string(template_name, context, request)

        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def upload_attendance(request):
    data = dict()
    template_name='attendance/upload_attendance.html'

    if request.is_ajax(): 
        if request.method == 'GET':
            form = CutOffPeriodForm(request.GET or None, request.FILES or None)
        elif request.method == 'POST':
            form = CutOffPeriodForm(request.POST or None, request.FILES or None) 

            if form.is_valid():
                try:
                    if 'attendance_file' in request.FILES:
                        att_file = request.FILES['attendance_file']
                        print(att_file.name) 
                        #https://code.djangoproject.com/ticket/26641
                        # employees = PersonalInfo.objects.values_list('id', flat=True).order_by('id')
                        #employees = PersonalInfo.objects.annotate(name=Concat('first_name', Value(' '),'middle_name', Value(' '),'last_name'),).filter(Q(name__icontains='lloyds Salazar Garcia'))
                        
                        # .query to raw sql query
                        #print(employees.query) 
                        #data['form_is_valid'] = True
                        read_attendance_file(att_file, request, form, data) 
                    else:
                        print('No File')                     
                except IntegrityError:
                    data['form_is_valid'] = False 
                    messages.error(request, 'Record Already Exists')
            else:
                data['form_is_valid'] = False
        context = {
            'form':form,
        }
        data['html_form'] =  render_to_string(template_name,context, request)
        return JsonResponse(data)
    else:
        raise Http404()

    # For Employees

@login_required
def reports_page(request):
    template_name = "reports/reports_page.html"
    user = get_object_or_404(User, username=request.user.username) 
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()
    if user.is_active and user.is_staff and user.is_superuser: 
        if request.method == 'GET':
         pass

        elif request.method == 'POST':
            pass
        context = {
            'user':user, 
            'notifications': notifications,
            'notifications_count': notifications_count, 
        }
        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def maintainance_page(request):
    template_name = "maintainance_page.html"
    user = get_object_or_404(User, username=request.user.username) 
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()
    if user.is_active and user.is_staff and user.is_superuser: 
        if request.method == 'GET':
            search_term = request.GET.get('search_term')
            if search_term.strip(): 
                print(search_term)

        elif request.method == 'POST':
            pass
        context = {
            'user':user, 
            'notifications': notifications,
            'notifications_count': notifications_count,
            'search_term': search_term,
        }
        return render(request, template_name, context)
    else:
        raise Http404()

# employee side-----------------------------------------------------------------------------

class PayrallAttendanceEmployeeIndexPage(LoginRequiredMixin, TemplateView):
    template_name = "employee_index_page.html"
    
    def get_context_data(self, *args, **kwargs):

        context = super(PayrallAttendanceEmployeeIndexPage,
                        self).get_context_data(*args, **kwargs)
        user = get_object_or_404(User, username=self.request.user.username)
        employee = get_object_or_404(PersonalInfo, fk_user=user)

        if user.is_active and user.is_staff and not user.is_superuser: 
            notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
            notifications_count = notifications.count()
            cutoffperiods = CutOffPeriodInfo.objects.all().distinct().count()
            total_leaves = EmployeeLeaves.objects.all().filter(Q(employee_leave_fk=employee)).distinct().count()  
            total_unapproved_leaves = EmployeeLeaves.objects.all().filter(Q(status='Disapproved') & Q(employee_leave_fk=employee)).distinct().count()  
            total_approved_leaves = EmployeeLeaves.objects.all().filter(Q(status='Approved')& Q(employee_leave_fk=employee)).distinct().count()  

            total_concern =  Concerns.objects.all().filter(Q(sender=employee)).distinct().count() 
            total_unreplied_concern = Concerns.objects.all().filter(Q(reply=None) & Q(sender=employee)).distinct().count() 
            
            total_unapproved_iteneraries = EmployeeItenerary.objects.all().filter(Q(~Q(checked_by=None) & ~Q(approved_by=None)) & Q(employee_itenerary_fk=employee)).distinct().count()  
            total_approved_iteneraries = EmployeeItenerary.objects.all().filter(Q(Q(checked_by=None) & Q(approved_by=None)) &  Q(employee_itenerary_fk=employee) ).distinct().count() 

            emp_payrolls = EmployeePayroll.objects.all().filter(Q(employee_fk=employee))
            year_list = []
            yearly_employee_salary_list = []
            yearly_emplpyee_deduction_list = []

            for emp in emp_payrolls:
                year = str(emp.payroll_cutoff_period)[:4]
                if not year in year_list:
                    year_list.append(year) 

            for year in year_list:
                total_net_pay = 0
                total_deductions = 0
                cutoffperiod = CutOffPeriodInfo.objects.all().filter(Q(cut_off_period__icontains=year)) 
                for cutoff in cutoffperiod:
                    emp_payroll = EmployeePayroll.objects.all().filter(Q(payroll_cutoff_period=cutoff) & Q(employee_fk=employee)).aggregate(Sum('net_pay')).get('net_pay__sum') 
                    emp_deductions  = EmployeePayroll.objects.all().filter(Q(payroll_cutoff_period=cutoff) & Q(employee_fk=employee)).aggregate(Sum('total_deduction')).get('total_deduction__sum') 
                    if emp_payroll:
                        total_net_pay = total_net_pay + emp_payroll 
                        total_deductions = total_deductions + emp_deductions
                yearly_employee_salary_list.append({'year': year, 'value': int(round(total_net_pay,2))})
                yearly_emplpyee_deduction_list.append({'year': year, 'deduction': int(round(total_deductions,2)), 'wages': int(round(total_net_pay,2))})
            
            
            dumps_yearly_employee_salary_list = json.dumps(yearly_employee_salary_list)
            loads_yearly_employee_salary_list =json.loads(dumps_yearly_employee_salary_list)
            # print(dumps_yearly_employee_wages_list)
            
            dumps_yearly_employee_deduction_list = json.dumps(yearly_emplpyee_deduction_list)
            loads_yearly_employee_deduction_list =json.loads(dumps_yearly_employee_deduction_list)

            
            context.update({
                'user': user,
                'cutoffperiods': cutoffperiods,
                'total_leaves':total_leaves,
                'total_unapproved_leaves': total_unapproved_leaves,
                'total_approved_leaves': total_approved_leaves,
                'total_concern': total_concern,
                'total_unreplied_concern': total_unreplied_concern,
                'total_unapproved_iteneraries':total_unapproved_iteneraries,
                'total_approved_iteneraries': total_approved_iteneraries,
                'loads_yearly_employee_salary_list': loads_yearly_employee_salary_list,
                'loads_yearly_employee_deduction_list':loads_yearly_employee_deduction_list, 
                'notifications': notifications,
                'notifications_count': notifications_count,
            })
            return context
        else:
            raise Http404()


@login_required
def employee_profile(request):
    template_name = "employee_side/employee_profile_page.html"
    user = get_object_or_404(User, username=request.user.username)
    last_time_logged = user.last_login
    last_time_joined = user.date_joined
    notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
    notifications_count = notifications.count()

    
    if user.is_active and user.is_staff and not user.is_superuser:
        if request.method == "GET": 
            record = PersonalInfo.objects.filter(fk_user=user)
            company = CompanyInfo.objects.filter(fk_company_user=user)
            profiles =  PersonalInfo.objects.all().filter(fk_user=user).distinct()
            company_details = CompanyInfo.objects.all().filter(fk_company_user=user).distinct()

        
        elif request.method == "POST":
            pass
    

        context = {
            'user': user,
            'last_time_logged': last_time_logged,
            'last_time_joined': last_time_joined, 
            'record': record,
            'profiles': profiles,
            'company':company,
            'company_details':company_details,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def side_employee_cut_off_payroll_page(request):
    template_name="employee_side/employee_side_cutoff_payroll_page.html"
    user = get_object_or_404(User, username=request.user.username)
    notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
    notifications_count = notifications.count()

    cut_off_list = CutOffPeriodInfo.objects.all().order_by('-id').distinct()

    if user.is_active and user.is_staff and not user.is_superuser:
        if request.method == 'GET':
            pass

        context = {
            'user': user,
            'cut_off_list': cut_off_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def side_employee_cut_off_page(request, id):
    template_name = "employee_side/employee_side_cutoff_page.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
    attendance_list = AttendanceInfo.objects.filter(Q(cut_off_period=cutoff) & Q(employee_profile=employee)).order_by("-id")
    notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
    notifications_count = notifications.count()
    

    if user.is_active and user.is_staff and not user.is_superuser:
        if request.method == 'GET':
            pass

        context = {
            'user': user, 
            'cutoff': cutoff,
            'attendance_list': attendance_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def side_employee_view_payroll_page(request, id):
    template_name = "employee_side/employee_side_view_payroll_page.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
    notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
    notifications_count = notifications.count()
 
    employee_salary = get_object_or_404(EmployeeSalary, employee_salary_fk=employee) 

    attendance_list = AttendanceInfo.objects.filter(Q(cut_off_period=cutoff) & Q(employee_profile=employee)).order_by("-id")
    try:
        emp_payroll_count = EmployeePayroll.objects.all().get(Q(employee_fk=employee) & Q(payroll_cutoff_period=cutoff))
    except EmployeePayroll.DoesNotExist:
        emp_payroll_count = EmployeePayroll.objects.all().filter(Q(employee_fk=employee) & Q(payroll_cutoff_period=cutoff)).distinct().count()
    today = date.today()
    date_today = today.strftime("%m/%d/%Y") 

    if user.is_active and user.is_staff and not user.is_superuser:
        if request.method == 'GET':
            pass

        context = {
            'user': user, 
            'cutoff': cutoff, 
            'attendance_list':attendance_list,
            'emp_payroll_count': emp_payroll_count,
            'employee_salary': employee_salary,
            'date_today': date_today,
            'employee': employee,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def side_employee_print_payroll_page(request, id):    
    template_name = "employee_side/employee_side_print_payroll_page.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    cutoff = get_object_or_404(CutOffPeriodInfo, id=id)
 
    employee_salary = get_object_or_404(EmployeeSalary, employee_salary_fk=employee) 

    attendance_list = AttendanceInfo.objects.filter(Q(cut_off_period=cutoff) & Q(employee_profile=employee)).order_by("-id")
    try:
        emp_payroll_count = EmployeePayroll.objects.all().get(Q(employee_fk=employee) & Q(payroll_cutoff_period=cutoff))
    except EmployeePayroll.DoesNotExist:
        emp_payroll_count = EmployeePayroll.objects.all().filter(Q(employee_fk=employee) & Q(payroll_cutoff_period=cutoff)).distinct().count()
    today = date.today()
    date_today = today.strftime("%m/%d/%Y") 

    if user.is_active and user.is_staff and not user.is_superuser:
        if request.method == 'GET':
            pass

        context = {
            'user': user, 
            'cutoff': cutoff, 
            'attendance_list':attendance_list,
            'emp_payroll_count': emp_payroll_count,
            'employee_salary': employee_salary,
            'date_today': date_today,
            'employee': employee,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def side_employee_manage_leaves_page(request):
    template_name = "employee_side/employee_side_leaves_page.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    leaves = EmployeeLeaves.objects.all().filter(employee_leave_fk=employee).order_by('-id').distinct()
    notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
    notifications_count = notifications.count()
    if user.is_active and user.is_staff and not user.is_superuser:
        if request.method == 'GET':
            pass


        context = {
            'user': user,  
            'employee': employee,
            'leaves': leaves,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def side_employee_create_leave_form(request):
    template_name = "employee_side/employee_side_create_leave.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)

    if user.is_active and user.is_staff and not user.is_superuser:
        if request.method == 'GET':
            form = EmployeeLeavesForm(request.GET or None)
        elif request.method == 'POST':
            form = EmployeeLeavesForm(request.POST or None)
            if form.is_valid():
                instance = form.save(commit=False) 
                instance.employee_leave_fk = employee
                instance.department = user.company_to_user.department
                instance.save()
                #lookup from parent table
                url = HttpResponseRedirect(reverse_lazy('application:employee_manage_leaves_page'))
                admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                for admin in admins:
                    Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="You have been received leave form from {sender}".format(sender=user),category=category_list[9],level=level_list[1])
                
                return HttpResponseRedirect(reverse_lazy('application:employee_side_manage_leaves_page'))
            else:
                print("Form Error!")

        context = {
            'user': user,  
            'employee': employee, 
            'form': form,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def side_employee_edit_leave_form(request, id):
    template_name = "employee_side/employee_side_edit_leave.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    leave = get_object_or_404(EmployeeLeaves, id=id)
    notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
    notifications_count = notifications.count()

    if user.is_active and user.is_staff and not user.is_superuser:
        if request.method == 'GET':
            form = EmployeeLeavesForm(request.GET or None, instance=leave)
        elif request.method == 'POST':
            form = EmployeeLeavesForm(request.POST or None, instance=leave)
            if form.is_valid():
                instance = form.save(commit=False)  
                instance.save()
                url = HttpResponseRedirect(reverse_lazy('application:employee_manage_leaves_page'))
                admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                for admin in admins:
                    Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="Leave form from {sender} has been updated".format(sender=user),category=category_list[10],level=level_list[1])
                
                return HttpResponseRedirect(reverse_lazy('application:employee_side_manage_leaves_page'))
            else:
                print("Form Error!")
        context = {
            'user': user,  
            'employee': employee, 
            'form': form,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def side_employee_delete_leave_form(request, id):
    data = dict()
    template_name = "employee_side/employee_side_delete_leave.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    leave = get_object_or_404(EmployeeLeaves, id=id)

    if request.is_ajax():
        if user.is_active and user.is_staff and not user.is_superuser:
            if request.method == 'GET':
                context = {
                'user': user,  
                'employee': employee,   
                'leave': leave,          
                } 
                data['html_form'] = render_to_string(template_name, context, request)
            elif request.method == 'POST':
                leave.delete()
                data['form_is_valid'] = True 
                url = HttpResponseRedirect(reverse_lazy('application:employee_manage_leaves_page'))
                admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                for admin in admins:
                    Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="Leave form from {sender} has been deleted!".format(sender=user),category=category_list[11],level=level_list[1])
                
            return JsonResponse(data)
        else:
            raise Http404()
    else:
        raise Http404()

@login_required
def side_employee_view_leave_form(request, id):
    template_name="employee_side/employee_side_view_leave.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    leave = get_object_or_404(EmployeeLeaves, id=id)
    notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
    notifications_count = notifications.count()

    if user.is_active and user.is_staff and not user.is_superuser:
        if request.method == 'GET':
            pass
        elif request.method == 'POST':
            pass

        context = {
            'user': user,  
            'employee': employee,  
            'leave': leave,
            'notifications': notifications,
            'notifications_count': notifications_count,

        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def side_employee_manage_inteneraries(request):
    template_name = "employee_side/employee_side_manage_iteneraries.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    itenerary_list = EmployeeItenerary.objects.all().filter(Q(employee_itenerary_fk=employee)).order_by('-id').distinct()
    notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
    notifications_count = notifications.count()

    if user.is_active and user.is_staff and not user.is_superuser:
        if request.method == 'GET':
            pass 
        context = {
            'user': user,  
            'employee': employee, 
            'itenerary_list': itenerary_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()


@login_required
def side_employee_create_inteneraries(request):
    template_name = "employee_side/employee_side_create_itenerary.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)

    if user.is_active and user.is_staff and not user.is_superuser:
        EmployeeIteneraryDetailsFormset = inlineformset_factory(EmployeeItenerary, EmployeeIteneraryDetails, form=EmployeeIteneraryDetailsForm, extra=1, can_delete=False) 
            
        if request.method == 'GET':
            formEID = EmployeeIteneraryDetailsFormset(request.GET or None)
            formIF = EmployeeIteneraryForm(request.GET or None)
        
        elif request.method == 'POST':
            formEID = EmployeeIteneraryDetailsFormset(request.POST or None)
            formIF = EmployeeIteneraryForm(request.POST or None)

            if formEID.is_valid() and formIF.is_valid(): 
                
                instanceFIF = formIF.save(commit=False)
                instanceFIF.employee_itenerary_fk = employee
                instanceFIF.save()

                instanceFEID = formEID.save(commit=False)

                for form in instanceFEID:
                    form.employee_itenerary = instanceFIF
                    form.save()

                #lookup from parent table
                url = HttpResponseRedirect(reverse_lazy('application:employee_manage_iteneraries_page'))
                admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                for admin in admins:
                    Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="You have been received itinerary form from {sender}".format(sender=user),category=category_list[9],level=level_list[1])
                

                return HttpResponseRedirect(reverse_lazy('application:employee_side_manage_iteneraries_page')) 
        context = { 
            'user': user,  
            'formIF': formIF,
            'employee': employee,  
            'formEID': formEID,
        }

        return render(request, template_name, context)
    else:
        raise Http404()


@login_required
def side_employee_edit_inteneraries(request, id):
    template_name = "employee_side/employee_side_edit_itenerary.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    itenerary = get_object_or_404(EmployeeItenerary, id=id)
    notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
    notifications_count = notifications.count()

    if user.is_active and user.is_staff and not user.is_superuser:
        #remove extra
        EmployeeIteneraryDetailsFormset = inlineformset_factory(EmployeeItenerary, EmployeeIteneraryDetails, form=EmployeeIteneraryDetailsForm, extra=0, can_delete=True) 
            
        if request.method == 'GET':
            formEID = EmployeeIteneraryDetailsFormset(request.GET or None, instance=itenerary) 
        
        elif request.method == 'POST':
            formEID = EmployeeIteneraryDetailsFormset(request.POST or None, instance=itenerary) 

            if formEID.is_valid():
                formEID.save()# you do not need to to loop through when deleting
                # instanceFEID = formEID.save(commit=False) 
                # for form in instanceFEID: 
                #     form.save() 
                 #lookup from parent table
                url = HttpResponseRedirect(reverse_lazy('application:employee_manage_iteneraries_page'))
                admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                for admin in admins:
                    Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="Itinerary form from {sender} has been updated!".format(sender=user),category=category_list[10],level=level_list[1])
               
                return HttpResponseRedirect(reverse_lazy('application:employee_side_manage_iteneraries_page')) 
        context = {
            'user': user,  
            'employee': employee, 
            'formEID': formEID,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()


@login_required
def side_employee_delete_inteneraries(request, id):
    data = dict()
    template_name = "employee_side/employee_side_delete_itenerary.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    itenerary = get_object_or_404(EmployeeItenerary, id=id)

    if request.is_ajax():
        if user.is_active and user.is_staff and not user.is_superuser:
            if request.method == 'GET': 
                context = {
                'user': user,  
                'employee': employee,   
                'itenerary': itenerary,          
                } 
                data['html_form'] = render_to_string(template_name, context, request)
            elif request.method == 'POST':
                itenerary.delete()
                data['form_is_valid'] = True 
                url = HttpResponseRedirect(reverse_lazy('application:employee_manage_iteneraries_page'))
                admins = PersonalInfo.objects.all().filter(Q(fk_user__is_superuser=True))
                for admin in admins:
                    Notifications.objects.create(sender=user,recipient=admin.fk_user,url=url.url,message="Itinerary form from {sender} has been deleted!".format(sender=user),category=category_list[11],level=level_list[1])
               
            return JsonResponse(data)
        else:
            raise Http404()
    else:
        raise Http404()


@login_required
def side_employee_view_itenerary_form(request, id):
    template_name = "employee_side/employee_side_view_itenerary.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    itenerary = get_object_or_404(EmployeeItenerary, id=id)
    itenerary_details = EmployeeIteneraryDetails.objects.all().filter(Q(employee_itenerary=itenerary)).order_by('-id').distinct()
    notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
    notifications_count = notifications.count()

    if user.is_active and user.is_staff and not user.is_superuser:
        if request.method == 'GET':
            pass
        elif request.method == 'POST':
            pass 
        context = {
            'user': user,  
            'employee': employee,  
            'itenerary': itenerary,
            'itenerary_details': itenerary_details,
            'notifications': notifications,
            'notifications_count': notifications_count,
        } 
        return render(request, template_name, context)
    else:
        raise Http404()


@login_required
def side_employee_manage_concerns(request):
    template_name = "employee_side/employee_side_manage_concerns.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
    notifications_count = notifications.count()


    concern_list = Concerns.objects.all().filter(Q(sender=employee)).order_by('-id').distinct()

    if user.is_active and user.is_staff and not user.is_superuser:
        if request.method == 'GET':
            pass 
        context = {
            'user': user,  
            'employee': employee, 
            'concern_list': concern_list,
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()

@login_required
def side_employee_create_concern(request):
    data = dict()
    template_name = "employee_side/employee_side_create_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)

    if user.is_active and user.is_staff and not user.is_superuser: 
        if request.method == 'GET':
            form = ConcernsEmployeeForm(request.GET or None)
        elif request.method == 'POST':
            form = ConcernsEmployeeForm(request.POST or None)   
            if form.is_valid():
                instance = form.save(commit=False)         
                instance.sender  = employee
                instance.save()

                date_time_str = str(instance.date_filed)
                date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d')
                converted_date = date_time_obj.strftime('%b %d, %Y')
                edit_url = reverse_lazy('application:employee_side_edit_concerns_page', kwargs={'id':instance.id})
                delete_url = reverse_lazy('application:employee_side_delete_concerns_page', kwargs={'id':instance.id})
                view_url = reverse_lazy('application:employee_side_view_concerns_page', kwargs={'id':instance.id})
                concern_date_dict = {
                    "id": instance.id,
                    "subject": instance.subject,
                    "date_filed": str(converted_date),
                    "edit_url": str(edit_url),
                    "view_url": str(view_url),
                    "delete_url": str(delete_url),
                }
                dict_to_string = json.dumps(concern_date_dict)
                data['concern_date_dict'] = json.loads(dict_to_string)
                data['form_is_valid'] = True
                if instance.receiver.fk_user.is_superuser:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_manage_inbox_concerns_page'))
                else:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_inbox_concerns_page'))
                Notifications.objects.create(sender=user,recipient=instance.receiver.fk_user,url=url.url,message="You had received new concern from {sender}".format(sender=employee),category=category_list[9],level=level_list[1])               

        context = {
            'form': form,
            'user': user,  
            'employee': employee,  
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()
    

@login_required
def side_employee_edit_concern(request, id):
    data = dict()
    template_name = "employee_side/employee_side_edit_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    concern = get_object_or_404(Concerns, id=id)
    
    if user.is_active and user.is_staff and not user.is_superuser: 
        if request.method == 'GET':
            form = ConcernsEmployeeForm(request.GET or None, instance=concern)
        elif request.method == 'POST':
            form = ConcernsEmployeeForm(request.POST or None, instance=concern)   
            if form.is_valid():
                instance = form.save(commit=False)   
                instance.save()

                date_time_str = str(instance.date_filed)
                date_time_obj = datetime.datetime.strptime(date_time_str, '%Y-%m-%d')
                converted_date = date_time_obj.strftime('%b %d, %Y')

                edit_url = reverse_lazy('application:employee_side_edit_concerns_page', kwargs={'id':concern.id})
                delete_url = reverse_lazy('application:employee_side_delete_concerns_page', kwargs={'id':concern.id})
                view_url = reverse_lazy('application:employee_side_view_concerns_page', kwargs={'id':concern.id})

                concern_date_dict = {
                    "id": instance.id,
                    "subject": instance.subject,
                    "date_filed": str(converted_date),
                    "edit_url": str(edit_url),
                    "view_url": str(view_url),
                    "delete_url": str(delete_url),
                }
                dict_to_string = json.dumps(concern_date_dict)
                data['concern_date_dict'] = json.loads(dict_to_string)
                data['form_is_valid'] = True

                if instance.receiver.fk_user.is_superuser:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_manage_inbox_concerns_page'))
                else:
                    url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_inbox_concerns_page'))
                Notifications.objects.create(sender=user,recipient=instance.receiver.fk_user,url=url.url,message="Concern from {sender} was updated".format(sender=employee),category=category_list[10],level=level_list[1])               


        context = {
            'form': form,
            'user': user,  
            'employee': employee,  
            'concern': concern,
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def side_employee_delete_concern(request, id):
    data = dict()
    template_name = "employee_side/employee_side_delete_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    concern = get_object_or_404(Concerns, id=id)
    
    if user.is_active and user.is_staff and not user.is_superuser: 
        if request.method == 'GET':
            context = { 
            'user': user,  
            'employee': employee,  
            'concern': concern,
            } 
            data['html_form'] = render_to_string(template_name, context, request)
        elif request.method == 'POST': 
            if concern.receiver.fk_user.is_superuser:
                url = HttpResponseRedirect(reverse_lazy('application:employee_manage_inbox_concerns_page'))
            else:
                url = HttpResponseRedirect(reverse_lazy('application:employee_side_manage_inbox_concerns_page')) 
            Notifications.objects.create(sender=user,recipient=concern.receiver.fk_user,url=url.url,message="Your concern from {sender}".format(sender=concern.sender),category=category_list[11],level=level_list[1])               
            concern.delete()
            data['form_is_valid'] = True 
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def side_employee_view_concern(request, id):
    data = dict()
    template_name = "employee_side/employee_side_view_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    concern = get_object_or_404(Concerns, id=id)
    
    if user.is_active and user.is_staff and not user.is_superuser: 
        if request.method == 'GET':
            pass
       
        context = { 
            'user': user,  
            'employee': employee,  
            'concern': concern,
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()


@login_required
def side_employee_manage_inbox_concern(request):
    template_name = "employee_side/employee_side_manage_inbox_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    notifications = Notifications.objects.all().filter(Q(recipient=user) | Q(public=True)).order_by('-id')
    notifications_count = notifications.count()

    concern_list = Concerns.objects.all().filter(Q(receiver=employee)).order_by('-id').distinct()

    if user.is_active and user.is_staff and not user.is_superuser:
        if request.method == 'GET':
            pass 
        context = {
            'user': user,  
            'employee': employee, 
            'concern_list': concern_list, 
            'notifications': notifications,
            'notifications_count': notifications_count,
        }

        return render(request, template_name, context)
    else:
        raise Http404()


@login_required
def side_employee_reply_concern(request, id):
    data = dict()
    template_name = "employee_side/employee_side_reply_concern.html"
    user = get_object_or_404(User, username=request.user.username)
    employee = get_object_or_404(PersonalInfo, fk_user=user)
    concern = get_object_or_404(Concerns, id=id)

    if user.is_active and user.is_staff and not user.is_superuser: 
        if request.method == 'GET':
            form = ConcernsReplyEmployeeForm(request.GET or None, instance=concern)
        elif request.method == 'POST':
            form = ConcernsReplyEmployeeForm(request.POST or None, instance=concern)   
            if form.is_valid():
                instance = form.save(commit=False)   
                instance.save()  

                data['form_is_valid'] = True

        context = {
            'form': form,
            'user': user,  
            'employee': employee,  
            'concern': concern,
        }

        data['html_form'] = render_to_string(template_name, context, request)
        return JsonResponse(data)
    else:
        raise Http404()

@login_required
def employee_side_maintainance_page(request):
    template_name = "employee_side/employee_side_maintainance_page.html"
    user = get_object_or_404(User, username=request.user.username) 
    notifications = Notifications.objects.all().filter(Q(Q(recipient=user) | Q(public=True)) & Q(is_read=False)).order_by('-id')
    notifications_count = notifications.count()
    if user.is_active and user.is_staff and not user.is_superuser: 
        if request.method == 'GET':
            search_term = request.GET.get('search_term')
            if search_term.strip(): 
                print(search_term)

        elif request.method == 'POST':
            pass
        context = {
            'user':user, 
            'notifications': notifications,
            'notifications_count': notifications_count,
            'search_term': search_term,
        }
        return render(request, template_name, context)
    else:
        raise Http404()
